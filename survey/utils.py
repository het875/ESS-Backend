import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def export_answers_to_excel(answers, file_path='answers.xlsx'):
    # Create a new workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Answer Sheet"

    # Add headers
    headers = ["Question", "Employee ID", "Answer", "Submitted At"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}1"] = header
        sheet[f"{col_letter}1"].font = Font(bold=True)

    # Add data rows
    for row_num, answer in enumerate(answers, start=2):
        sheet[f"A{row_num}"] = answer["question_text"]
        sheet[f"B{row_num}"] = answer["employee_id"]
        sheet[f"C{row_num}"] = answer["answer_text"]
        sheet[f"D{row_num}"] = answer["submitted_at"]

    # Adjust column widths
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet.column_dimensions[col_letter].width = 20

    # Save the workbook
    workbook.save(file_path)
    print(f"Excel file has been created: {file_path}")
